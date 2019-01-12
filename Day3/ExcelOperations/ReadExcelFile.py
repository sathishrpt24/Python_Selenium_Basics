import openpyxl as xl
print("How are you")

test = xl.load_workbook("C:\\Users\\acer\\Desktop\\demo.xlsx")
#Sheetnames = test.get_sheet_names()

workSheet = test.get_sheet_by_name("Sheet1")

value = workSheet.max_row
value1 = workSheet.max_column

print(workSheet['A2'].value)
print("Last active Row",value)
print("Last active column",value1)

for row in range(1,20):
    for column in range (1,5):
        print(workSheet.cell(row,column).value,"\t",end='')
        #print('\t')
    print('\n')
test.close()